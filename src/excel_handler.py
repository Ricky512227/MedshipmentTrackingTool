"""
Excel Handler Module for MedshipmentTrackingTool

This module handles all Excel file operations including reading input data,
writing output files, and generating categorized reports.
"""

import os
import json
import datetime
import unicodedata
from openpyxl import load_workbook
import xlsxwriter


class ExcelHandler:
    """Handles all Excel file operations for the tracking tool."""
    
    def __init__(self, config_path="config/config.json"):
        """
        Initialize ExcelHandler with configuration.
        
        Args:
            config_path (str): Path to configuration file
        """
        with open(config_path, 'r') as f:
            self.config = json.load(f)
        
        # Ensure output directories exist
        os.makedirs(self.config['output_dir'], exist_ok=True)
        os.makedirs(self.config['items_dir'], exist_ok=True)
    
    def read_input_data(self):
        """
        Reads tracking data from the input Excel file.
        
        Returns:
            tuple: (order_ids, first_names, last_names, tracking_numbers)
        """
        try:
            wb = load_workbook(self.config['input_file'])
            ws = wb.active
            
            # Read columns A, B, C, D
            order_column = ws['A']
            firstName_column = ws['B']
            lastName_column = ws['C']
            trackingNum_column = ws['D']
            
            # Extract values (skip header row)
            order_ids = [cell.value for cell in order_column[1:] if cell.value]
            first_names = [cell.value for cell in firstName_column[1:] if cell.value]
            last_names = [cell.value for cell in lastName_column[1:] if cell.value]
            tracking_numbers = [cell.value for cell in trackingNum_column[1:] if cell.value]
            
            print(f"Collected {len(tracking_numbers)} tracking numbers from Excel")
            return order_ids, first_names, last_names, tracking_numbers
            
        except FileNotFoundError:
            print(f"Error: Input file not found at {self.config['input_file']}")
            raise
        except Exception as e:
            print(f"Error reading input file: {e}")
            raise
    
    def write_final_data(self, tracking_data):
        """
        Writes the complete tracking data to Final-Data.xlsx.
        
        Args:
            tracking_data (list): List of tracking records to write
        """
        workbook = xlsxwriter.Workbook(self.config['final_data_file'])
        worksheet = workbook.add_worksheet()
        
        # Define headers
        headers = [
            'Local Date and Time', 'Country', 'Location', 'OrderId',
            'First Name', 'Last Name', 'Tracking Number',
            'Event Type', 'Mail Category', 'Next Office', 'Extra Information'
        ]
        
        # Write headers
        bold_format = workbook.add_format({'bold': True})
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, bold_format)
        
        # Write data
        for row, record in enumerate(tracking_data, start=1):
            for col, value in enumerate(record):
                worksheet.write(row, col, value)
        
        workbook.close()
        print(f"Final data written to {self.config['final_data_file']}")
    
    def categorize_shipments(self):
        """
        Reads Final-Data.xlsx and categorizes shipments by status.
        
        Returns:
            dict: Dictionary mapping status names to row numbers
        """
        wb = load_workbook(self.config['final_data_file'])
        ws = wb.active
        event_column = ws['H']  # Event Type column
        
        # Define event types and their corresponding status names
        event_mappings = {
            "Receive item from customer (Otb)": "Booked",
            "Receive item at office of exchange (Otb)": "Booked",
            "Insert item into bag (Otb)": "InTransit",
            "Receive item at office of exchange (Inb)": "InTransit",
            "Receive item at delivery office (Inb)": "InTransitToDelivery",
            "Deliver item (Inb)": "Delivered",
            "Send item to customs (Inb)": "InBound",
            "Return item from customs (Inb)": "OutBound",
            "Unsuccessful item delivery attempt (Inb)": "NoticeLeft",
            "Receive item at collection point for pick-up (Inb)": "NoticeLeft",
            "Send item to domestic location (Inb)": "Returned",
            "Record item customs information (Inb)": "Stuck"
        }
        
        # Initialize category lists
        categories = {
            "Booked": [],
            "InTransit": [],
            "InBound": [],
            "OutBound": [],
            "Delivered": [],
            "NoticeLeft": [],
            "InTransitToDelivery": [],
            "Stuck": [],
            "Returned": []
        }
        
        # Categorize rows based on event type
        for idx, cell in enumerate(event_column[1:], start=2):  # Start from row 2 (skip header)
            if cell.value in event_mappings:
                category = event_mappings[cell.value]
                categories[category].append(idx)
        
        return categories
    
    def extract_row_numbers(self, cell_list):
        """
        Extracts row numbers from openpyxl cell objects.
        
        Args:
            cell_list (list): List of openpyxl cell objects
            
        Returns:
            list: List of row numbers (integers)
        """
        row_numbers = []
        for cell in cell_list:
            try:
                # Extract row number from cell object string representation
                cell_str = str(cell)
                row_num = int(cell_str.split(".")[1].replace(">", "").replace("H", ""))
                row_numbers.append(row_num)
            except (ValueError, IndexError) as e:
                print(f"Error extracting row number from {cell}: {e}")
                continue
        return row_numbers
    
    def generate_categorized_report(self, categories):
        """
        Generates Excel file with categorized shipments in separate sheets.
        
        Args:
            categories (dict): Dictionary mapping category names to row numbers
            
        Returns:
            str: Path to generated file
        """
        # Load the final data workbook
        wb = load_workbook(self.config['final_data_file'])
        ws = wb.active
        
        # Generate timestamp for filename
        timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        output_file = os.path.join(self.config['items_dir'], f'Data{timestamp}.xlsx')
        
        workbook = xlsxwriter.Workbook(output_file)
        
        # Generate sheet for each category
        for category_name, row_numbers in categories.items():
            if len(row_numbers) > 0:
                self._create_category_sheet(workbook, ws, category_name, row_numbers)
        
        # Create summary sheet
        self._create_summary_sheet(workbook, categories)
        
        workbook.close()
        print(f"Categorized report generated: {output_file}")
        return output_file
    
    def _create_category_sheet(self, workbook, worksheet_parent, sheet_name, row_numbers):
        """
        Creates a worksheet for a specific category.
        
        Args:
            workbook: xlsxwriter workbook object
            worksheet_parent: openpyxl worksheet to read from
            sheet_name (str): Name of the sheet to create
            row_numbers (list): List of row numbers to include
        """
        worksheet = workbook.add_worksheet(sheet_name)
        
        # Write headers
        headers = [
            'Local Date and Time', 'Country', 'Location', 'OrderId',
            'First Name', 'Last Name', 'Tracking Number',
            'Event Type', 'Mail Category', 'Next Office', 'Extra Information'
        ]
        bold_format = workbook.add_format({'bold': True})
        
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, bold_format)
        
        # Write data rows
        for row_idx, row_num in enumerate(row_numbers, start=1):
            try:
                row_data = list(worksheet_parent.rows)[row_num - 1]
                values = []
                
                for cell in row_data:
                    if cell.value is None:
                        values.append("NotFound")
                    else:
                        # Normalize unicode characters
                        try:
                            value = unicodedata.normalize('NFKD', str(cell.value)).encode('ascii', 'ignore').decode('ascii')
                            values.append(value)
                        except:
                            values.append(str(cell.value))
                
                # Ensure we have the right number of columns
                while len(values) < len(headers):
                    values.append("")
                
                for col_idx, value in enumerate(values[:len(headers)]):
                    worksheet.write(row_idx, col_idx, value)
                    
            except Exception as e:
                print(f"Error writing row {row_num} to {sheet_name}: {e}")
                continue
    
    def _create_summary_sheet(self, workbook, categories):
        """
        Creates a summary sheet with counts for each category.
        
        Args:
            workbook: xlsxwriter workbook object
            categories (dict): Dictionary mapping category names to row numbers
        """
        worksheet = workbook.add_worksheet("Summary")
        
        # Write headers
        worksheet.write(0, 0, "ITEMS", workbook.add_format({'bold': True}))
        worksheet.write(0, 1, "COUNT", workbook.add_format({'bold': True}))
        
        # Calculate totals
        total_items = sum(len(rows) for rows in categories.values())
        
        # Write category counts
        summary_data = [
            ("Delivered Items", len(categories.get("Delivered", []))),
            ("Booked Items", len(categories.get("Booked", []))),
            ("Items in InBound", len(categories.get("InBound", []))),
            ("Items in Intransit", len(categories.get("InTransit", []))),
            ("Items in Notice Left", len(categories.get("NoticeLeft", []))),
            ("Items in OutBound", len(categories.get("OutBound", []))),
            ("Intransit to Delivery items", len(categories.get("InTransitToDelivery", []))),
            ("Items in Stuck", len(categories.get("Stuck", []))),
            ("Return Items", len(categories.get("Returned", []))),
            ("TOTAL ITEMS", total_items)
        ]
        
        for row_idx, (item_name, count) in enumerate(summary_data, start=1):
            worksheet.write(row_idx, 0, item_name)
            worksheet.write(row_idx, 1, count)
